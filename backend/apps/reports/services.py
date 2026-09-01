import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from django.db.models import Count
from django.utils import timezone
from django.contrib.auth import get_user_model

from apps.attendance.models import AttendanceRecord, AttendanceSession
from apps.academics.models import Class, Section
from apps.accounts.models import StudentProfile

User = get_user_model()


class DashboardService:
    @staticmethod
    def get_dashboard_summary(user):
        today = timezone.localdate()

        total_students = User.objects.filter(role=User.Role.STUDENT, is_active=True).count()
        total_teachers = User.objects.filter(role=User.Role.TEACHER, is_active=True).count()
        total_classes = Class.objects.count()

        face_registered = StudentProfile.objects.filter(is_registration_complete=True).count()
        face_pending = StudentProfile.objects.filter(is_registration_complete=False).count()

        today_sessions = AttendanceSession.objects.filter(date=today)
        today_records = AttendanceRecord.objects.filter(session__date=today)

        total_marked_today = today_records.count()
        present_today = today_records.filter(status=AttendanceRecord.Status.PRESENT).count()
        absent_today = today_records.filter(status=AttendanceRecord.Status.ABSENT).count()
        late_today = today_records.filter(status=AttendanceRecord.Status.LATE).count()

        attendance_rate = (
            round((present_today / total_marked_today) * 100, 1)
            if total_marked_today > 0
            else 0.0
        )

        class_analytics = list(
            Class.objects.annotate(student_count=Count("students")).values("id", "name", "student_count")
        )

        # Class-wise attendance for today (present / marked / total students).
        class_attendance = []
        for cls in Class.objects.all():
            class_students = cls.students.count()
            class_records = AttendanceRecord.objects.filter(
                session__class_obj=cls, session__date=today
            )
            class_present = class_records.filter(status=AttendanceRecord.Status.PRESENT).count()
            class_attendance.append(
                {
                    "class_id": cls.id,
                    "class_name": cls.name,
                    "student_count": class_students,
                    "present": class_present,
                    "marked": class_records.count(),
                }
            )

        section_analytics = list(
            Section.objects.select_related("class_obj").annotate(student_count=Count("students")).values(
                "id", "name", "class_obj__name", "student_count"
            )
        )

        recent_records = (
            AttendanceRecord.objects.select_related("student", "session", "session__class_obj")
            .order_by("-marked_at")[:10]
        )

        recent_list = [
            {
                "id": r.id,
                "student_name": r.student.full_name,
                "class_name": r.session.class_obj.name if r.session and r.session.class_obj else "N/A",
                "status": r.status,
                "method": r.verification_method,
                "marked_at": r.marked_at.strftime("%Y-%m-%d %H:%M:%S"),
            }
            for r in recent_records
        ]

        return {
            "total_students": total_students,
            "total_teachers": total_teachers,
            "total_classes": total_classes,
            "face_registered": face_registered,
            "face_pending": face_pending,
            "today_sessions": today_sessions.count(),
            "today_attendance": total_marked_today,
            "today_marked": total_marked_today,
            "today_present": present_today,
            "today_absent": absent_today,
            "today_late": late_today,
            "attendance_rate": attendance_rate,
            "class_analytics": class_analytics,
            "class_attendance": class_attendance,
            "section_analytics": section_analytics,
            "recent_activity": recent_list,
        }


class ReportExportService:
    @staticmethod
    def generate_excel_report(queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Headers styling
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        align_center = Alignment(horizontal="center", vertical="center")

        headers = [
            "Date",
            "Student Roll No",
            "Student Name",
            "Class",
            "Section",
            "Status",
            "Verification Method",
            "Confidence Score",
            "Time Marked",
        ]
        ws.append(headers)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

        for rec in queryset:
            student_profile = getattr(rec.student, "student_profile", None)
            roll = student_profile.roll_number if student_profile else ""
            c_name = rec.session.class_obj.name if rec.session and rec.session.class_obj else ""
            s_name = rec.session.section_obj.name if rec.session and rec.session.section_obj else ""

            ws.append(
                [
                    str(rec.session.date) if rec.session else "",
                    roll,
                    rec.student.full_name,
                    c_name,
                    s_name,
                    rec.status,
                    rec.verification_method,
                    rec.confidence_score or "N/A",
                    rec.marked_at.strftime("%H:%M:%S"),
                ]
            )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def generate_pdf_report(queryset):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=colors.HexColor("#1E3A8A"),
            spaceAfter=12,
        )
        subtitle_style = ParagraphStyle(
            "ReportSub",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.HexColor("#4B5563"),
            spaceAfter=20,
        )

        elements.append(Paragraph("AI Face Attendance Report", title_style))
        elements.append(
            Paragraph(
                f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {queryset.count()}",
                subtitle_style,
            )
        )

        table_data = [["Date", "Roll No", "Student Name", "Class", "Status", "Method"]]

        for rec in queryset[:500]:
            student_profile = getattr(rec.student, "student_profile", None)
            roll = student_profile.roll_number if student_profile else ""
            c_name = rec.session.class_obj.name if rec.session and rec.session.class_obj else ""

            table_data.append(
                [
                    str(rec.session.date) if rec.session else "",
                    roll,
                    rec.student.full_name[:20],
                    c_name,
                    rec.status,
                    rec.verification_method,
                ]
            )

        t = Table(table_data, colWidths=[70, 70, 150, 80, 70, 100])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F9FAFB")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),
                ]
            )
        )
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
